import argparse
import os
from Colors import Colors
from openpyxl import load_workbook
import selenium
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome import service
import Levenshtein
from tabulate import tabulate


driverPath = '/Users/star-yoshi/plaac/chromedriver'
chromiumPath = '/Users/star-yoshi/plaac/chromedriver'
urlSYGDS288c = 'https://nribf1.nrib.go.jp/SYGD/search.cgi?prj=01303&sgnm=1&sobj=gen'
urlSYGDK7 = 'https://nribf1.nrib.go.jp/SYGD/search.cgi?prj=01303&sgnm=2&sobj=gen'


def getAA(driver,geneName):
    driver.get(urlSYGDS288c + '&sgen_freetxt=' + geneName)
    res = driver.find_elements(By.XPATH,"//div[@id='tab1']/div/table[@class='listing']/tbody/tr[@class]")
    #strlist = [Colors.CYAN + geneName + Colors.RESET]
    #ratlist = [None]
    
    if res:
        hitres = res[0].find_element(By.XPATH,"./td[position()=3]")
        ratmax = Levenshtein.ratio(geneName,hitres.text)
    
        for r in res[1:]:
            hit = r.find_element(By.XPATH,"./td[position()=3]")
            rat = Levenshtein.ratio(geneName,hit.text)
            if rat > ratmax:
                ratmax = rat
                hitres = hit
        
        hitname = hitres.text
        link=hitres.find_element(By.XPATH,"./preceding-sibling::td/a").get_attribute('href')
        driver.get(link)
        tabs=driver.find_element(By.XPATH,"//div[@class='ui-tabs ui-corner-all ui-widget ui-widget-content']")
        tab3=tabs.find_element(By.XPATH,"./ul[@role='tablist']/li[position()=3]")
        tab3.click()
        aa = tabs.find_elements(By.XPATH,"./div[position()=3]/div/dt/textarea")
        if  len(aa) > 0:
            aa = aa[0].text
        else:
            print(Colors.CYAN + geneName + Colors.RESET)
            return None
        
        print(
            Colors.CYAN + geneName + Colors.RESET + ' : ' + 
            Colors.MAGENTA + hitname + Colors.RESET + ' : ' + 
            Colors.YELLOW + link + Colors.RESET
        )
        print(aa)
        return aa[0] + hitname + '|' + aa[1:]
    else:
        print(Colors.CYAN + geneName + Colors.RESET)
        return None

parser = argparse.ArgumentParser()
parser.add_argument('inputFile',help='input file')
parser.add_argument('outputFile',help='output file')

args = parser.parse_args()

os.system('')

book = load_workbook(args.inputFile)
sheet = book[book.sheetnames[0]]

cellIn = sheet.cell(2,4)
cellOut = sheet.cell(2,12)

geneList = []

cell = cellIn
while cell.value is not None: 
    #print(cell.value)
    geneList.append(cell.value)
    cell = cell.offset(1,0)

chrome_service = service.Service(executable_path=driverPath)
options = ChromeOptions()
options.add_argument('--headless')
options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
#options.binary_location = chromiumPath
driver = selenium.webdriver.Chrome(service=chrome_service,options=options)

aalist = []

try:    
    for g in geneList:
        aa = getAA(driver,g)
        if aa:
            aalist.append(aa)

finally:
    driver.quit()

if args.outputFile:
    with open(args.outputFile,'w') as f:
        f.write('\n'.join(aalist))